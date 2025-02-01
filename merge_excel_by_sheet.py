import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import xlwings as xw
import os
import threading
from mask_layer import show_overlay

def is_sheet_empty(sheet):
    """判断工作表是否为空"""
    for row in sheet.iter_rows(min_row=1):  # 从第一行开始检查
        for cell in row:
            if cell.value is not None:
                return False
    return True


def merge_excel_files(source_folder, target_file, root):
    # 封装好的遮罩层
    overlay = show_overlay(root)
    try:
        # 创建一个新的 Excel 文件
        merged_wb = openpyxl.Workbook()
        # 删除默认的工作表
        del merged_wb[merged_wb.sheetnames[0]]

        for root_dir, dirs, files in os.walk(source_folder):
            for file in files:
                if file.endswith(('.xls', '.xlsx')):
                    file_path = os.path.join(root_dir, file)
                    if file.endswith('.xlsx'):
                        # 处理 .xlsx 文件
                        wb = openpyxl.load_workbook(file_path)
                        try:
                            for sheet_name in wb.sheetnames:
                                source_sheet = wb[sheet_name]
                                if is_sheet_empty(source_sheet):
                                    continue  # 如果工作表为空，跳过该工作表

                                # 创建新的工作表，名称与原工作表一致
                                target_sheet = merged_wb.create_sheet(title=sheet_name)

                                # 复制列宽
                                for col_letter, dimension in source_sheet.column_dimensions.items():
                                    target_sheet.column_dimensions[col_letter].width = dimension.width

                                # 复制单元格内容和格式
                                for row in source_sheet.iter_rows():
                                    for cell in row:
                                        target_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                                        target_cell.font = cell.font.copy()
                                        target_cell.border = cell.border.copy()
                                        target_cell.fill = cell.fill.copy()
                                        target_cell.number_format = cell.number_format
                                        target_cell.protection = cell.protection.copy()
                                        target_cell.alignment = cell.alignment.copy()
                                # 复制合并单元格
                                for merged_cell in source_sheet.merged_cells.ranges:
                                    target_sheet.merge_cells(str(merged_cell))
                        except Exception as e:
                            # 显示错误提示框
                            root.after(0, lambda: messagebox.showerror("错误", f".xlsx格式合并过程中出现错误：{str(e)}"))
                        finally:
                            wb.close()
                    else:
                        # 处理 .xls 文件
                        app = xw.App(visible=False)
                        wb = app.books.open(file_path)
                        try:
                            for sheet in wb.sheets:
                                # 将 xlwings 的 sheet 转换为 openpyxl 可操作的临时文件
                                new_wb = app.books.add()
                                sheet.copy(before=new_wb.sheets[0])
                                new_wb.save(f"{os.path.splitext(target_file)[0]}_{sheet.name}.xlsx")
                                new_wb.close()

                                temp_wb = openpyxl.load_workbook(f"{os.path.splitext(target_file)[0]}_{sheet.name}.xlsx")
                                temp_sheet = temp_wb.active
                                if is_sheet_empty(temp_sheet):
                                    temp_wb.close()
                                    os.remove(f"{os.path.splitext(target_file)[0]}_{sheet.name}.xlsx")
                                    continue  # 如果工作表为空，跳过该工作表

                                # 创建新的工作表，名称与原工作表一致
                                target_sheet = merged_wb.create_sheet(title=sheet.name)

                                # 复制列宽
                                for col_letter, dimension in temp_sheet.column_dimensions.items():
                                    target_sheet.column_dimensions[col_letter].width = dimension.width

                                for row in temp_sheet.iter_rows():
                                    for cell in row:
                                        target_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                                        target_cell.font = cell.font.copy()
                                        target_cell.border = cell.border.copy()
                                        target_cell.fill = cell.fill.copy()
                                        target_cell.number_format = cell.number_format
                                        target_cell.protection = cell.protection.copy()
                                        target_cell.alignment = cell.alignment.copy()
                                for merged_cell in temp_sheet.merged_cells.ranges:
                                    target_sheet.merge_cells(str(merged_cell))
                                temp_wb.close()
                                os.remove(f"{os.path.splitext(target_file)[0]}_{sheet.name}.xlsx")
                        except Exception as e:
                            # 显示错误提示框
                            root.after(0, lambda: messagebox.showerror("错误", f".xls格式合并过程中出现错误：{str(e)}"))
                        finally:
                            wb.close()
                            app.quit()
        # 保存合并后的文件
        merged_wb.save(target_file)

        root.after(0, lambda: messagebox.showinfo("完成", "Excel 文件按页签合并完成！"))
    except Exception as e:
        # 显示错误提示框
        root.after(0, lambda: messagebox.showerror("错误", f"合并过程中出现错误：{str(e)}"))
    finally:
        # 销毁遮罩
        overlay.destroy()

def show_merge_excel_by_sheet(frame, root):

    def select_source_folder():
        folder = filedialog.askdirectory()
        if folder:
            source_entry.delete(0, tk.END)
            source_entry.insert(0, folder)

    def select_target_file():
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file:
            target_entry.delete(0, tk.END)
            target_entry.insert(0, file)

    def merge_excel():
        excel_path = source_entry.get()
        output_dir = target_entry.get()
        if not excel_path or not output_dir:
            messagebox.showwarning("提示", "请选择源文件夹和目标文件！")
            return
        threading.Thread(target=merge_excel_files, args=(excel_path, output_dir, root)).start()

    # 源文件夹输入框
    source_label = tk.Label(frame, text="源文件夹:")
    source_label.grid(row=0, column=0, padx=5, pady=5, sticky=tk.E)
    excel_label = tk.Label(frame, text="*", fg="red")
    excel_label.grid(row=0, column=1, padx=5, pady=5, sticky=tk.E)
    source_entry = tk.Entry(frame, width=35)
    source_entry.grid(row=0, column=2, padx=5, pady=5, sticky=tk.EW)
    source_button = tk.Button(frame, text="浏览", command=select_source_folder)
    source_button.grid(row=0, column=3, padx=5, pady=5, sticky=tk.W)

    # 目标文件输入框
    target_label = tk.Label(frame, text="目标文件:")
    target_label.grid(row=1, column=0, padx=5, pady=5, sticky=tk.E)
    excel_label = tk.Label(frame, text="*", fg="red")
    excel_label.grid(row=1, column=1, padx=5, pady=5, sticky=tk.E)
    target_entry = tk.Entry(frame, width=35)
    target_entry.grid(row=1, column=2, padx=5, pady=5, sticky=tk.EW)
    target_button = tk.Button(frame, text="浏览", command=select_target_file)
    target_button.grid(row=1, column=3, padx=5, pady=5, sticky=tk.W)

    # 合并按钮
    merge_button = tk.Button(frame, text="合并Excel文件", command=merge_excel)
    merge_button.grid(row=2, column=2, pady=10,sticky=tk.EW)

    # 功能概述
    label_text = ("注：该功能为按页签将多个Excel文件进行合并,支持后缀为.xls和.xlsx两种格式（建议使用.xlsx为后缀的文件），"
                  "输出合并后的文件统一为.xlsx后缀，输出文件的页签和源文件页签命名保持一致。浏览选择需要合并的目录，"
                  "且选定保存的目标Excel，执行合并即可。")
    output_label = tk.Label(frame, text=label_text, justify=tk.LEFT, wraplength=627)
    output_label.grid(row=3, column=0, padx=5, pady=20, columnspan=4)