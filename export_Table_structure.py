import tkinter as tk
from tkinter import messagebox, filedialog
import dmPython
import pandas as pd
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import threading
from mask_layer import show_overlay

# 用于保存用户选择的文件路径
selected_file_path = None


def export_to_excel(server_val, port_val, username_val, password_val, schema_val, table_name_val, selected_file_path, root):

    # 封装好的遮罩层
    overlay = show_overlay(root)

    try:

        try:
            conn = dmPython.connect(user=username_val, password=password_val, server=server_val, port=port_val)
            cursor = conn.cursor()
            try:
                # 查询数据
                sql = """
                        SELECT 
                            -- 表名
                            ut.table_name,
                            -- 字段名
                            uc.column_name,
                            -- 数据类型
                            uc.data_type,
                            -- 长度
                            CASE 
                                WHEN uc.data_type IN ('CHAR', 'VARCHAR', 'BINARY', 'VARBINARY') THEN uc.char_length
                                WHEN uc.data_type IN ('NCHAR', 'NVARCHAR2') THEN uc.char_length
                                WHEN uc.data_type IN ('NUMBER') THEN uc.data_precision
                                ELSE NULL
                            END AS data_length,
                            -- 精度
                            uc.data_scale,
                            -- 是否主键
                            CASE 
                                WHEN upc.column_name IS NOT NULL THEN '是'
                                ELSE '否'
                            END AS is_primary_key,
                            -- 备注
                            ucc.comments
                        FROM 
                            -- 用户表信息视图
                            dba_tables ut
                        JOIN 
                            -- 用户列信息视图
                            dba_tab_columns uc ON ut.table_name = uc.table_name AND ut.owner = uc.owner
                        LEFT JOIN 
                            -- 用户主键约束列信息视图
                            (
                                SELECT 
                                    ucc.table_name,
                                    ucc.column_name
                                FROM 
                                    dba_constraints uc
                                JOIN 
                                    dba_cons_columns ucc ON uc.constraint_name = ucc.constraint_name AND uc.owner = ucc.owner
                                WHERE 
                                    uc.constraint_type = 'P'
                            ) upc ON ut.table_name = upc.table_name AND uc.column_name = upc.column_name
                        LEFT JOIN 
                            -- 用户列注释信息视图
                            dba_col_comments ucc ON ut.table_name = ucc.table_name AND uc.column_name = ucc.column_name AND ut.owner = ucc.owner
                        WHERE 
                            -- 指定模式名，将 ? 替换为实际的模式名
                            ut.owner = ?
                    """
                # 添加模式参数
                params = []
                params.append(schema_val)
                # 判断是否输入了表名
                if table_name_val.strip() == "":
                    sql += " ORDER BY ut.table_name, uc.column_id"
                else:
                    sql += " AND ut.table_name = ? ORDER BY ut.table_name, uc.column_id"
                    # 添加表名参数
                    params.append(table_name_val)

                print(sql)
                cursor.execute(sql, tuple(params))
                res = cursor.fetchall()

                # 定义表头
                headers = ['表名', '字段名', '数据类型', '长度', '精度', '是否主键', '备注']

                # 将查询结果转换为 DataFrame
                df = pd.DataFrame(res, columns=headers)

                # 创建 Excel 文件
                wb = Workbook()
                ws = wb.active

                # 写入表头
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    # 设置表头样式
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    border = Border(left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))
                    cell.border = border
                    # 设置表头颜色
                    cell.fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")

                # 写入数据
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        # 设置数据边框
                        border = Border(left=Side(style='thin'),
                                        right=Side(style='thin'),
                                        top=Side(style='thin'),
                                        bottom=Side(style='thin'))
                        cell.border = border
                        # 设置数据垂直居中
                        cell.alignment = Alignment(vertical='center')

                # 设置行高和列高
                for row in ws.iter_rows(min_row=1, max_row=df.shape[0] + 1):
                    for cell in row:
                        ws.row_dimensions[cell.row].height = 20
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 20

                # 保存 Excel 文件
                wb.save(selected_file_path)
                messagebox.showinfo("导出成功", "Excel 文件导出成功！")

            except (dmPython.Error, Exception) as err:
                messagebox.showerror("导出失败", f"Excel 文件导出失败：{err}")
            finally:
                cursor.close()
                conn.close()
        except (dmPython.Error, Exception) as err:
            messagebox.showerror("连接失败", f"数据库连接失败：{err}")

    except Exception as e:
        # 显示错误提示框
        root.after(0, lambda: messagebox.showerror("错误", f"合并过程中出现错误：{str(e)}"))
    finally:
        # 销毁遮罩
        overlay.destroy()

def show_export_Table_structure(frame, root):

    def test_connection():
        server_val = server_entry.get()
        port_val = port_entry.get()
        username_val = username_entry.get()
        password_val = password_entry.get()
        try:
            conn = dmPython.connect(user=username_val, password=password_val, server=server_val, port=port_val)
            conn.close()
            messagebox.showinfo("连接成功", "数据库连接测试成功！")
        except (dmPython.Error, Exception) as err:
            messagebox.showerror("连接失败", f"数据库连接测试失败：{err}")

    def select_save_path():
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file:
            save_path_entry.delete(0, tk.END)
            save_path_entry.insert(0, file)

    def start_export():
        server_val = server_entry.get()
        port_val = port_entry.get()
        username_val = username_entry.get()
        password_val = password_entry.get()
        schema_val = schema_entry.get()
        table_name_val = table_name_entry.get()
        selected_file_path = save_path_entry.get()

        if not selected_file_path:
            messagebox.showwarning("提示", "请选择源文件夹和目标文件！")
            return
        threading.Thread(target=export_to_excel, args=(server_val, port_val, username_val, password_val, schema_val, table_name_val, selected_file_path, root)).start()


    # ==============================================================================

    # 创建输入框和标签
    tk.Label(frame, text="主机名:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.E)
    tk.Label(frame, text="*", fg="red").grid(row=0, column=1, padx=5, pady=5, sticky=tk.E)
    server_entry = tk.Entry(frame, width=35)
    server_entry.grid(row=0, column=2, padx=5, pady=5, sticky=tk.EW)

    tk.Label(frame, text="端口:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.E)
    tk.Label(frame, text="*", fg="red").grid(row=1, column=1, padx=5, pady=5, sticky=tk.E)
    port_entry = tk.Entry(frame, width=35)
    port_entry.grid(row=1, column=2, padx=5, pady=5, sticky=tk.EW)

    tk.Label(frame, text="用户名:").grid(row=2, column=0, padx=5, pady=5, sticky=tk.E)
    tk.Label(frame, text="*", fg="red").grid(row=2, column=1, padx=5, pady=5, sticky=tk.E)
    username_entry = tk.Entry(frame, width=35)
    username_entry.grid(row=2, column=2, padx=5, pady=5, sticky=tk.EW)

    tk.Label(frame, text="口令:").grid(row=3, column=0, padx=5, pady=5, sticky=tk.E)
    tk.Label(frame, text="*", fg="red").grid(row=3, column=1, padx=5, pady=5, sticky=tk.E)
    password_entry = tk.Entry(frame, show="*", width=35)
    password_entry.grid(row=3, column=2, padx=5, pady=5, sticky=tk.EW)

    tk.Label(frame, text="模式:").grid(row=4, column=0, padx=5, pady=5, sticky=tk.E)
    tk.Label(frame, text="*", fg="red").grid(row=4, column=1, padx=5, pady=5, sticky=tk.E)
    schema_entry = tk.Entry(frame, width=35)
    schema_entry.grid(row=4, column=2, padx=5, pady=5, sticky=tk.EW)

    tk.Label(frame, text="表名:").grid(row=5, column=0, padx=5, pady=5, sticky=tk.E)
    # tk.Label(frame, text="*", fg="red").grid(row=5, column=1, padx=5, pady=5, sticky=tk.E)
    table_name_entry = tk.Entry(frame, width=35)
    table_name_entry.grid(row=5, column=2, padx=5, pady=5, sticky=tk.EW)

    # 新增保存路径输入框和按钮
    tk.Label(frame, text="保存文件:").grid(row=6, column=0, padx=5, pady=5, sticky=tk.E)
    tk.Label(frame, text="*", fg="red").grid(row=6, column=1, padx=5, pady=5, sticky=tk.E)
    save_path_entry = tk.Entry(frame, width=35)
    save_path_entry.grid(row=6, column=2, padx=5, pady=5, sticky=tk.EW)
    select_path_button = tk.Button(frame, text="浏览", command=select_save_path)
    select_path_button.grid(row=6, column=3, padx=5, pady=5, sticky=tk.W)

    # 创建按钮
    test_button = tk.Button(frame, text="测试连接", command=test_connection)
    test_button.grid(row=7, column=0, pady=10)

    export_button = tk.Button(frame, text="导出 Excel 表结构", command=start_export)
    export_button.grid(row=7, column=2, pady=10, sticky=tk.EW)


    # 功能概述
    label_text = ("注：目前只支持达梦数据库，可以将选定模式下所有表或单个表的表结构输出为Excel文件。")
    output_label = tk.Label(frame, text=label_text, justify=tk.LEFT, wraplength=627)
    output_label.grid(row=8, column=0, padx=5, pady=20, columnspan=4)