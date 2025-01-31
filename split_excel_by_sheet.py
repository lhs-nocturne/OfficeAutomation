import tkinter as tk
from tkinter import filedialog, messagebox
import os
import threading
from openpyxl import load_workbook
import xlwings as xw
from mask_layer import show_overlay

# ============================拆分xls格式的方法 =================================
def split_xls_by_sheet(excel_path, output_dir, root, overlay):
    # 确保输出文件夹存在，如果不存在则创建
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 启动 Excel 应用程序
    app = xw.App(visible=False)
    # 打开输入的 Excel 文件
    wb = app.books.open(excel_path)
    try:
        # 遍历每个工作表
        for sheet in wb.sheets:
            # 创建一个新的工作簿
            new_wb = app.books.add()

            # 复制当前工作表到新工作簿
            sheet.copy(before=new_wb.sheets[0])

            # 删除新工作簿的默认工作表
            new_wb.sheets[1].delete()

            # 生成输出文件路径
            output_file = os.path.join(output_dir, f"{sheet.name}.xls")
            # 保存新工作簿
            new_wb.save(output_file)
            # 关闭新工作簿
            new_wb.close()
            print(f"已保存工作表 {sheet.name} 到 {output_file}")

    except Exception as e:
        messagebox.showerror(f"处理过程中出现错误: {e}")
    finally:
        # 关闭输入的工作簿
        if 'wb' in locals():
            wb.close()
        # 退出 Excel 应用程序
        app.quit()
        # 关闭遮罩
        overlay.destroy()

# ==========================拆分xls格式的方法 =================================
def split_xlsx_by_sheet(excel_path, output_dir, root, overlay):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 加载 Excel 文件
    wb = load_workbook(excel_path)
    try:
        # 遍历每个工作表
        for sheet_name in wb.sheetnames:
            # 创建一个新的工作簿
            new_wb = load_workbook(excel_path)
            # 获取当前工作表
            current_sheet = new_wb[sheet_name]

            # 删除其他工作表
            for other_sheet_name in new_wb.sheetnames:
                if other_sheet_name != sheet_name:
                    other_sheet = new_wb[other_sheet_name]
                    new_wb.remove(other_sheet)

            # 构建输出文件路径
            output_file = os.path.join(output_dir, f"{sheet_name}.xlsx")

            # 保存新的工作簿
            new_wb.save(output_file)
            print(f"已保存工作表 {sheet_name} 到 {output_file}")
    except Exception as e:
        root.after(0, lambda: messagebox.showerror("错误", f"拆分过程中出现错误：{str(e)}"))
    finally:
        # 关闭源工作簿
        wb.close()
        # 关闭遮罩
        overlay.destroy()


def split_excel_worker(excel_path, output_dir, root):

    # 封装好的遮罩层
    overlay = show_overlay(root)
    # 检查输出文件夹是否存在，不存在则创建
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    # 根据文件扩展名选择合适的处理方法
    if excel_path.endswith('.xls'):
        split_xls_by_sheet(excel_path, output_dir, root, overlay)
    elif excel_path.endswith('.xlsx'):
        split_xlsx_by_sheet(excel_path, output_dir, root, overlay)
    else:
        root.after(0, lambda: messagebox.showinfo("提示", "不支持的文件格式，请提供 .xls 或 .xlsx 文件。"))

    root.after(0, lambda: messagebox.showinfo("完成", "Excel 文件按页签拆分完成！"))


def show_split_excel_by_sheet(frame, root):

    def select_excel_file():
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if file_path:
            excel_entry.delete(0, tk.END)
            excel_entry.insert(0, file_path)

    def select_output_directory():
        dir_path = filedialog.askdirectory()
        if dir_path:
            output_entry.delete(0, tk.END)
            output_entry.insert(0, dir_path)

    def split_excel():
        excel_path = excel_entry.get()
        output_dir = output_entry.get()
        if not excel_path or not output_dir:
            messagebox.showwarning("提示", "请选择 Excel 文件和输出目录！")
            return
        threading.Thread(target=split_excel_worker, args=(excel_path, output_dir, root)).start()

    # 创建选择 Excel 文件的组件
    excel_label = tk.Label(frame, text="Excel文件:")
    excel_label.grid(row=0, column=0, padx=5, pady=5, sticky=tk.E)
    excel_entry = tk.Entry(frame, width=35)
    excel_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
    excel_button = tk.Button(frame, text="浏览", command=select_excel_file)
    excel_button.grid(row=0, column=2, padx=5, pady=5, sticky=tk.W)

    # 创建选择输出目录的组件
    output_label = tk.Label(frame, text="目标目录:")
    output_label.grid(row=1, column=0, padx=5, pady=5, sticky=tk.E)
    output_entry = tk.Entry(frame, width=35)
    output_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.EW)
    output_button = tk.Button(frame, text="浏览", command=select_output_directory)
    output_button.grid(row=1, column=2, padx=5, pady=5, sticky=tk.W)

    # 创建拆分按钮
    split_button = tk.Button(frame, text="拆分Excel文件", command=split_excel)
    split_button.grid(row=2, column=1, pady=10, sticky=tk.EW)

    # 功能概述
    label_text = "注：该功能为按页签拆分Excel文件，支持后缀为.xls和.xlsx两种格式，并且输出文件与输入文件后缀保持一致。浏览选择单个Excel文件，且选定拆分后的目标目录后，执行拆分即可。"
    output_label = tk.Label(frame, text=label_text, justify=tk.LEFT, wraplength=627)
    output_label.grid(row=3, column=0, padx=5, pady=20,  columnspan=3)
